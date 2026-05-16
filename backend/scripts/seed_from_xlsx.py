"""external.xlsx → /data/indicators/YYYY-MM-DD.csv + indicators 테이블.

MIH 자동 수집의 시뮬레이션입니다 (IPO § 2-0, § 2-10 P-09).

처리 흐름:
    1. external.xlsx 의 시트1(메인 데이터)만 사용
    2. 4개 제품의 feature_mapping(primary + aliases) 에 매칭되는 행만 필터
    3. CSV 로 저장: /data/indicators/YYYY-MM-DD.csv (수집 기준일별)
    4. indicators 테이블 upsert
"""
from __future__ import annotations

import asyncio
import csv
from collections.abc import Iterable
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert

from app.config import get_settings
from app.db import SessionLocal
from app.models import Indicator, Product

SETTINGS = get_settings()
SHEET_NAME = "시트1"


def _parse_date(value: Any, _cycle: str) -> date | None:
    """xlsx 의 수집날짜를 date 로 정규화.

    xlsx 원본은 같은 컬럼에 datetime, ISO 문자열, "YYYYMMDD" / "YYYYMM" / "YYYY" 숫자형
    이 섞여 있고, 일부 row 는 잘못된 월(13~) 데이터를 포함한다. 검증 실패 시 None 반환 →
    호출자가 해당 row 를 skip.
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    if s.endswith(".0"):
        s = s[:-2]
    try:
        return datetime.fromisoformat(s).date()
    except ValueError:
        pass
    if not s.isdigit():
        return None
    try:
        if len(s) == 8:
            return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
        if len(s) == 6:
            return date(int(s[:4]), int(s[4:6]), 1)
        if len(s) == 4:
            return date(int(s), 1, 1)
    except ValueError:
        return None
    return None


def _iter_xlsx_rows() -> Iterable[dict[str, Any]]:
    wb = load_workbook(SETTINGS.external_xlsx_path, data_only=True, read_only=True)
    ws = wb[SHEET_NAME]
    headers: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            headers = [str(h or "").strip() for h in row]
            continue
        if not any(row):
            continue
        record = dict(zip(headers, row, strict=False))
        cycle = str(record.get("수집주기(Y/M/W/D)") or "D").strip()
        date_value = _parse_date(record.get("수집날짜(기준)"), cycle)
        if date_value is None or record.get("값") is None:
            continue
        try:
            numeric = float(record["값"])
        except (TypeError, ValueError):
            continue
        yield {
            "source": (record.get("출처") or "").strip() or "Unknown",
            "country": (record.get("국가") or "").strip() or None,
            "category_big": (record.get("큰 카테고리") or "").strip() or None,
            "category_mid": (record.get("중간 카테고리") or "").strip() or None,
            "category_small": (record.get("소 카테고리") or "").strip() or None,
            "feature_name": (record.get("지표명") or "").strip(),
            "date": date_value,
            "value": numeric,
            "unit": (record.get("단위") or "").strip() or "-",
            "cycle": cycle.upper(),
        }


def _csv_dump(rows: list[dict[str, Any]]) -> int:
    """수집일별 CSV 분할 저장(IPO § 2-0)."""
    SETTINGS.indicators_csv_dir.mkdir(parents=True, exist_ok=True)
    buckets: dict[date, list[dict[str, Any]]] = {}
    for r in rows:
        buckets.setdefault(r["date"], []).append(r)
    fieldnames = ["source", "country", "category_big", "category_mid", "category_small", "feature_name", "date", "value", "unit", "cycle"]
    for d, items in buckets.items():
        path: Path = SETTINGS.indicators_csv_dir / f"{d.isoformat()}.csv"
        with path.open("w", encoding="utf-8", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(items)
    return len(buckets)


async def _load_feature_filter() -> set[str]:
    """PRD 0514 — products.key_features + axis 의 모든 지표를 평탄화."""
    async with SessionLocal() as s:
        products = (await s.execute(select(Product))).scalars().all()
    features: set[str] = set()
    for p in products:
        for f in p.key_features or []:
            features.add(f)
        for axis_grp in (p.axis or {}).values():
            for ind in axis_grp.get("indicators", []) or []:
                features.add(ind)
    return features


def _dedupe(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """동일 (feature_name, date, source) 키는 마지막 값으로 정규화.

    xlsx 원본은 같은 키에 여러 행이 존재하며, ON CONFLICT 구문은 같은 batch 내
    중복 키를 두 번 건드릴 수 없어 INSERT 가 실패한다. 입력단에서 한 번 정리.
    """
    seen: dict[tuple[str, date, str], dict[str, Any]] = {}
    for r in rows:
        seen[(r["feature_name"], r["date"], r["source"])] = r
    return list(seen.values())


async def _upsert_indicators(rows: list[dict[str, Any]]) -> int:
    if not rows:
        return 0
    rows = _dedupe(rows)
    async with SessionLocal() as s:
        stmt = insert(Indicator).values(rows).on_conflict_do_update(
            index_elements=["feature_name", "date", "source"],
            set_={
                "value": insert(Indicator).excluded.value,
                "unit": insert(Indicator).excluded.unit,
                "cycle": insert(Indicator).excluded.cycle,
                "country": insert(Indicator).excluded.country,
                "category_big": insert(Indicator).excluded.category_big,
                "category_mid": insert(Indicator).excluded.category_mid,
                "category_small": insert(Indicator).excluded.category_small,
            },
        )
        await s.execute(stmt)
        await s.commit()
    return len(rows)


async def main() -> None:
    feature_filter = await _load_feature_filter()
    if not feature_filter:
        raise SystemExit("products 가 비어있습니다. 먼저 seed_customers.py 를 실행하세요.")

    print(f"feature_filter ({len(feature_filter)} indicators):")
    for f in sorted(feature_filter):
        print(f"  - {f}")

    matched = [r for r in _iter_xlsx_rows() if r["feature_name"] in feature_filter]
    print(f"\nxlsx → 필터링된 행: {len(matched)}")

    csv_files = _csv_dump(matched)
    print(f"CSV 저장: {csv_files} 개 파일 in {SETTINGS.indicators_csv_dir}")

    inserted = await _upsert_indicators(matched)
    print(f"indicators upsert: {inserted} 행")


if __name__ == "__main__":
    asyncio.run(main())
